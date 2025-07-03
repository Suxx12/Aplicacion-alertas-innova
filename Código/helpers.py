from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import pandas as pd
import os

def generate_files(send_separately, executives, df_base_filtering, EXCEL_FOLDER, prefix, files, wanted_columns, type):
    if send_separately:
        for executive in executives:
            print(f"Generando archivo para: {executive}")
            if type == "T3" or type == "T4" or type == "T6": #agregamos el nuevo tipo T6
                df_executive = df_base_filtering[df_base_filtering['Ejecutivo Técnico Proyecto'] == executive]
            else:
                df_executive = df_base_filtering[df_base_filtering['Nombre Ejecutivo Técnico'] == executive]
                
            if not df_executive.empty:
                filename = os.path.join(EXCEL_FOLDER, f"{prefix}_{executive.replace(' ', '_')}.xlsx")
                save_format_excel(df_executive, filename, wanted_columns, type)
                files.append(filename)
            else:
                print(f"No hay datos para el ejecutivo: {executive}")
    else:
        print("Generando archivo consolidado para todos los ejecutivos seleccionados.")
        if type == "T3" or type == "T4" or type == "T6":  #agregamos el nuevo tipo T6
            df_consolidated = df_base_filtering[df_base_filtering['Ejecutivo Técnico Proyecto'].isin(executives)]
        else:
            df_consolidated = df_base_filtering[df_base_filtering['Nombre Ejecutivo Técnico'].isin(executives)]

        if not df_consolidated.empty:
            filename = os.path.join(EXCEL_FOLDER, f"{prefix}.xlsx")
            save_format_excel(df_consolidated, filename, wanted_columns, type)
            files.append(filename)
        else:
            print("No hay datos para generar el archivo consolidado.")

    print(f"Archivos generados: {files}")
   

def date_columns(df, columns):
    for col in columns:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce')
    return df

def save_format_excel(df, filename, wanted_columns, type):
    df_to_save = df[wanted_columns]

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    ws.append(wanted_columns)

    for row in dataframe_to_rows(df_to_save, index=False, header=False):
        ws.append(row)

    date_style_name = "date_style"
    if date_style_name not in wb.named_styles:
        date_style = NamedStyle(name=date_style_name, number_format="DD/MM/YYYY")
        wb.add_named_style(date_style)


    for row in ws.iter_rows(min_row=2, max_col=len(wanted_columns)):
        for cell in row:
            if type == "T1":
                if cell.column in [wanted_columns.index('Anticipo') + 1, wanted_columns.index('Fiel cumplimiento') + 1]:
                    cell.style = date_style
            elif type == "T3":
                if cell.column in [4, 5]:  
                    cell.style = date_style     
            elif type == "T4":
                if cell.column in [wanted_columns.index('Fecha Entrega Programada') + 1]:
                    cell.style = date_style         
            elif type == "T6":  # Nuevo tipo T6 agregado
                if cell.column in [wanted_columns.index('Fecha Entrega Programada') + 1, wanted_columns.index('Fecha Entrega Real') + 1]:
                    cell.style = date_style
            else:
                if cell.column in [wanted_columns.index('Fecha Resolucion')+ 1]:
                    cell.style = date_style

    date_columns = Table(displayName="date_columnsReporte", ref=f"A1:{get_column_letter(len(wanted_columns))}{len(df_to_save) + 1}")
    table_style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    date_columns.tableStyleInfo = table_style
    ws.add_table(date_columns)
    
    for col_idx, col_name in enumerate(wanted_columns, start=1):
        max_length = max(len(str(cell.value)) for cell in ws[get_column_letter(col_idx)] if cell.value)
        adjusted_width = max_length + 2  
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    wb.save(filename)

def reset_filters(df_original):
    return df_original.copy()  
